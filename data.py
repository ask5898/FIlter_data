import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter
from pandas import ExcelFile
from itertools import groupby,repeat 
import numpy as np

df = load_workbook(filename = "data.xlsx",read_only=True)
status_on = []
status_off = []
temp = []
thresh = 3
for x in df :
	for y in df.iter_rows() :
		for cell in y :
			print cell.value
#status_and_time = zip(df["Status"],df["TimeStamp"])
'''for status,time in status_and_time  :
	time = str(time)
	status = int(status)
	if status == 1 :
		status_on.append(time)
	else :
		status_off.append(time)

convert_to_list = lambda stat : [x for x in map(int,stat.split(":"))]

off = [convert_to_list(x) for x in status_off]
on = [convert_to_list(x) for x in status_on]

del temp,status_on,status_off
#segreg = {0 : off,
#         1 : on}
	
#temp_on = np.asarray(segreg[1] , dtype=np.int32)

off.sort()
on.sort()

length_on = [(key,len(list(group))) for key, group in groupby(on)]
length_off = [(key,len(list(group))) for key, group in groupby(off)]

for l_on,k in length_on :
	if k < thresh :
		on.remove(l_on)

for l_off,k in length_off :
	if k < thresh :
		off.remove(l_off)


def remove_duplicate_in_list(list) :
	final_list = []
	for num in list :
		if num not in final_list :
			final_list.append(num)

	return final_list

def remove_duplicate_between_list(list_n,list_m) :
	for x in list_n :
		if x in list_m :
			list_n.remove(x)
			list_m.remove(x)

on = remove_duplicate_in_list(on)
off = remove_duplicate_in_list(off)
remove_duplicate_between_list(on,off)

def close_time(list) :
	final_time = []
	for i in range(len(list)-1) :
		for j in range(i+1,len(list)-1) :
			if list[i][0] == list[j][0] :
				final_val_min = (list[i][1]+list[j][1])/2
				final_val_sec = (list[i][2]+list[j][2])/2
				final_time.append(list[i][0])
				final_time.append(final_val_min)
				final_time.append(final_val_sec)		
				list.pop(i)
				list.pop(j-1)
				list.append(final_time)
				final_time = []


close_time(on)
close_time(off)
on.sort()
off.sort()


while len(on) != len(off) :
	if len(on) > len(off) :
		on.pop()
	else :
		off.pop


final_on = final_off = []
msg_time = lambda st : ':'.join(map(str,st))
final_on = [(msg_time(i)) for i in on]
final_off = [(msg_time(i)) for i in off]
print final_on
print final_off'''

				
		





















		
